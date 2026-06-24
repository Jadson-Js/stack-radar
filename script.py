import os
import requests
import pandas as pd
import re
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# ==========================================
# 1. Configuration and Credentials
# ==========================================
APP_ID  = os.getenv('APP_ID')
APP_KEY = os.getenv('APP_KEY')

DAYS_AGO       = 365
MAX_WORKERS    = 2
RETRY_ATTEMPTS = 3
RETRY_DELAY    = 5
JOB_LIMIT      = 150_000

# Safety locks for dynamic pagination
MAX_PAGES_PER_TERM = 150  # Prevents infinite requests if anti-loop fails
MAX_CONSECUTIVE_FAILURES = 3  # Tolerates up to 3 consecutive empty/failed pages
MAX_CONSECUTIVE_SATURATION = 2 # Tolerates up to 2 full pages with 100% duplicate jobs

COUNTRIES: dict[str, dict] = {
    'br': {
        'name': 'Brazil',
        'terms': [
            'desenvolvedor', 'developer', 'programador', 'engenheiro de software',
            'software engineer', 'web developer', 'analista de sistemas',
            'pessoa desenvolvedora', 'dev', 'desenvolvedor fullstack',
            'desenvolvedor backend', 'desenvolvedor frontend', 'desenvolvedor mobile',
            'full stack developer', 'backend developer', 'frontend developer',
            'arquiteto de software', 'software architect', 'tech lead',
            'líder técnico', 'analista de TI', 'analista desenvolvedor',
        ],
    },
    'us': {
        'name': 'USA',
        'terms': [
            'developer', 'software engineer', 'web developer', 'application developer',
            'computer programmer', 'backend developer', 'backend engineer',
            'frontend developer', 'frontend engineer', 'full stack developer',
            'full stack engineer', 'mobile developer', 'ios developer',
            'android developer', 'software architect', 'solutions architect',
            'tech lead', 'engineering manager',
        ],
    },
}

# ==========================================
# 2. Level Classification
# ==========================================
_SENIOR_RE = re.compile(
    r's[eê]nior|sr\.?\s|pleno|mid[\s\-]?level|staff\s+engineer|lead\s+engineer|'
    r'principal\s+engineer|architect|tech\s+lead|engineering\s+manager',
    re.IGNORECASE
)
_JUNIOR_RE = re.compile(
    r'j[uú]nior|jr\.?\s|entry[\s\-]?level|iniciante|0[\s\-]?[aà][\s\-]?2\s*anos?|'
    r'early[\s\-]?career|associate\s+engineer|associate\s+developer',
    re.IGNORECASE
)
_INTERNSHIP_RE = re.compile(
    r'est[aá]gi[oá]rio?|est[aá]gio|intern(ship)?|trainee|aprendiz|co[\s\-]?op',
    re.IGNORECASE
)

ALL_LEVELS = ('Junior', 'Internship', 'Mid-level', 'Senior', 'General')

def classify_level(text: str) -> str:
    if _INTERNSHIP_RE.search(text):
        return 'Internship'
    if _JUNIOR_RE.search(text):
        return 'Junior'
    if _SENIOR_RE.search(text):
        if re.search(r'pleno|mid[\s\-]?level', text, re.IGNORECASE):
            return 'Mid-level'
        return 'Senior'
    return 'General'

# ==========================================
# 3. Categories and Aliases Dictionary
# ==========================================
CATEGORIES = {
    'Languages': [
        'JavaScript', 'Python', 'Java', 'C#', 'TypeScript', 'PHP', 'Ruby',
        'Golang', 'Go', 'Rust', 'Kotlin', 'Swift', 'C++', 'C', 'SQL', 'Dart',
        'Scala', 'R', 'COBOL', 'Perl', 'Elixir', 'Haskell', 'Lua', 'Shell',
        'Bash', 'PowerShell', 'Groovy', 'Clojure', 'F#', 'Objective-C',
        'Assembly', 'MATLAB', 'Julia',
    ],
    'Web Frameworks & Libs': [
        'React', 'Angular', 'Vue.js', 'Vue', 'Next.js', 'Nuxt.js', 'Nuxt',
        'Svelte', 'SvelteKit', 'Gatsby', 'Remix', 'Astro', 'Solid.js',
        'jQuery', 'Bootstrap', 'Tailwind', 'Tailwind CSS', 'Material UI',
        'Chakra UI', 'Ant Design', 'shadcn', 'Storybook',
    ],
    'Backend Frameworks & Libs': [
        'Node.js', 'Express', 'NestJS', 'Fastify', 'Koa',
        'Spring Boot', 'Spring', 'Spring MVC', 'Spring Security', 'Spring Cloud',
        'Django', 'Flask', 'FastAPI', 'Celery', 'SQLAlchemy',
        'Laravel', 'Symfony', 'CodeIgniter', 'Lumen',
        '.NET', 'ASP.NET', 'Entity Framework', 'Blazor',
        'Rails', 'Ruby on Rails', 'Sinatra',
        'Gin', 'Fiber', 'Echo', 'Actix', 'Rocket',
        'Ktor', 'Micronaut', 'Quarkus', 'Phoenix', 'Elixir Phoenix',
        'gRPC', 'GraphQL', 'REST', 'RESTful',
    ],
    'Mobile': [
        'Flutter', 'React Native', 'Ionic', 'Xamarin', 'MAUI',
        'Android', 'iOS', 'Swift', 'SwiftUI', 'Jetpack Compose',
        'Expo', 'Capacitor', 'Cordova',
    ],
    'Data & AI': [
        'TensorFlow', 'PyTorch', 'Keras', 'scikit-learn', 'Pandas',
        'NumPy', 'Spark', 'Apache Spark', 'Kafka', 'Apache Kafka',
        'Airflow', 'Apache Airflow', 'dbt', 'Hadoop', 'Hive',
        'Power BI', 'Tableau', 'Looker', 'Metabase', 'Superset',
        'MLflow', 'Langchain', 'LangChain', 'OpenAI', 'LLM', 'RAG',
        'Databricks', 'Snowflake', 'BigQuery', 'Redshift',
        'Machine Learning', 'Deep Learning', 'NLP', 'Computer Vision',
        'Jupyter', 'Matplotlib', 'Seaborn', 'Plotly',
    ],
    'Databases': [
        'PostgreSQL', 'MySQL', 'MongoDB', 'Redis', 'Oracle', 'SQL Server',
        'MariaDB', 'SQLite', 'Cassandra', 'DynamoDB', 'Elasticsearch',
        'Neo4j', 'CouchDB', 'InfluxDB', 'TimescaleDB', 'Supabase',
        'Firebase', 'PlanetScale', 'Neon', 'Cockroachdb',
    ],
    'DevOps & Cloud': [
        'Docker', 'Kubernetes', 'Helm', 'Istio', 'ArgoCD',
        'AWS', 'Azure', 'GCP', 'Google Cloud', 'DigitalOcean', 'Heroku', 'Vercel',
        'CI/CD', 'GitHub Actions', 'GitLab CI', 'Jenkins', 'CircleCI', 'Travis CI',
        'Terraform', 'Ansible', 'Puppet', 'Chef', 'Pulumi',
        'Linux', 'Unix', 'Nginx', 'Apache', 'Caddy',
        'Prometheus', 'Grafana', 'Datadog', 'Sentry', 'New Relic', 'Splunk',
        'CloudFormation', 'CDK', 'SAM', 'Lambda', 'ECS', 'EKS', 'S3', 'RDS', 'EC2',
        'Azure DevOps', 'Azure Kubernetes', 'AKS', 'GKE',
    ],
    'Tools & Practices': [
        'Git', 'GitHub', 'GitLab', 'Bitbucket', 'Jira', 'Confluence', 'Trello', 'Notion',
        'Scrum', 'Agile', 'Kanban', 'SAFe', 'XP', 'TDD', 'BDD', 'DDD', 'SOLID', 
        'Clean Architecture', 'Clean Code', 'Microservices', 'Microsserviços', 
        'Monorepo', 'Serverless', 'OpenAPI', 'Swagger', 'Postman', 'Insomnia',
        'WebSockets', 'WebSocket', 'MQTT', 'RabbitMQ', 'ActiveMQ', 'SQS',
        'OAuth', 'JWT', 'OpenID', 'SAML', 'SSO',
        'OWASP', 'Cybersecurity', 'Segurança', 'Penetration Testing',
        'Linux', 'Bash', 'Shell Script',
    ],
}

STACK_TO_CAT = {stack: cat for cat, stacks in CATEGORIES.items() for stack in stacks}

STACK_ALIASES: dict[str, str] = {
    'JavaScript': 'JavaScript/TypeScript',
    'TypeScript': 'JavaScript/TypeScript',
}

for _alias, _canonical in STACK_ALIASES.items():
    if _canonical not in STACK_TO_CAT and _alias in STACK_TO_CAT:
        STACK_TO_CAT[_canonical] = STACK_TO_CAT[_alias]

# ==========================================
# 4. Regex Compilation
# ==========================================
def _compile_regex(keyword: str) -> re.Pattern:
    escaped = re.escape(keyword)
    if ' ' in keyword:
        escaped = r'\s+'.join(re.escape(p) for p in keyword.split())
    return re.compile(rf'(?i)(?<![a-z0-9]){escaped}(?![a-z0-9])')

REGEX_MAP: dict[str, re.Pattern] = {
    stack: _compile_regex(stack)
    for stack in STACK_TO_CAT
    if stack not in STACK_ALIASES.values()
}

_REMOTE_RE = re.compile(r'remoto|home[\s\-]?office|remote|distributed|anywhere', re.IGNORECASE)
_HYBRID_RE = re.compile(r'híbrido|híbrida|hybrid', re.IGNORECASE)

# ==========================================
# 5. Defensive Search Engine
# ==========================================
def _fetch_page(country: str, term: str, page: int) -> list[dict] | None:
    """Returns a list of jobs. Returns None ONLY if the request fails completely."""
    url = f"https://api.adzuna.com/v1/api/jobs/{country}/search/{page}"
    params = {
        'app_id': APP_ID, 'app_key': APP_KEY, 'what': term,
        'max_days_old': DAYS_AGO, 'results_per_page': 50,
    }
    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            resp = requests.get(url, params=params, timeout=15)
            if resp.status_code in (429, 503):
                wait = RETRY_DELAY * attempt
                print(f"  ! [{country.upper()}] p{page:>3} — {resp.status_code}, waiting {wait}s")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            results = resp.json().get('results', [])
            print(f"  * [{country.upper()}] '{term[:20]:<20}' p{page:>3} — {len(results)} jobs")
            return results
        except Exception as exc:
            print(f"  x [{country.upper()}] p{page:>3} — error: {exc}")
            if attempt < RETRY_ATTEMPTS:
                time.sleep(RETRY_DELAY)
    return None

def get_country_jobs(country: str, terms: list[str], global_jobs: dict[str, dict]) -> dict[str, dict]:
    country_jobs: dict[str, dict] = {}

    for term in terms:
        if len(global_jobs) >= JOB_LIMIT:
            print(f"\n  [!] Global limit reached. Stopping collection.")
            break

        print(f"\n  > Searching '{term}' (Safe Pagination)...")
        page = 1
        consecutive_failures = 0
        consecutive_saturation = 0
        term_interrupted = False

        while page <= MAX_PAGES_PER_TERM:
            if len(global_jobs) >= JOB_LIMIT:
                term_interrupted = True
                break

            page_batch = list(range(page, page + MAX_WORKERS))
            batch_results: dict[int, list[dict] | None] = {}

            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                futures = {pool.submit(_fetch_page, country, term, p): p for p in page_batch}
                for fut in as_completed(futures):
                    batch_results[futures[fut]] = fut.result()

            stop_term = False
            for p in page_batch:
                results = batch_results.get(p)

                # Handling 1: API Error (None) or actually empty page ([])
                if not results: 
                    consecutive_failures += 1
                    if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                        print(f"  - '{term}' — {consecutive_failures} consecutive failures/empty pages. Terminating term.")
                        stop_term = True
                        break
                    continue
                else:
                    consecutive_failures = 0 # Reset failure counter

                # Handling 2: Saturation and global duplicates
                new_on_this_page = 0
                for job in results:
                    jid = job.get('id') or job.get('redirect_url', '')
                    if jid:
                        if jid not in global_jobs:
                            global_jobs[jid] = job
                            new_on_this_page += 1
                        # Add to country even if it's a global duplicate (e.g., BR and US sharing)
                        if jid not in country_jobs:
                            country_jobs[jid] = job

                if new_on_this_page == 0:
                    consecutive_saturation += 1
                    if consecutive_saturation >= MAX_CONSECUTIVE_SATURATION:
                        print(f"  - '{term}' — API Loop detected (only duplicate jobs). Terminating term.")
                        stop_term = True
                        break
                else:
                    consecutive_saturation = 0

                if len(global_jobs) >= JOB_LIMIT:
                    stop_term = True
                    term_interrupted = True
                    break

            if stop_term or term_interrupted:
                break
            page += MAX_WORKERS

        if page > MAX_PAGES_PER_TERM:
            print(f"  - '{term}' reached the safety limit ({MAX_PAGES_PER_TERM} pages).")

        print(f"  -> Unique jobs in country: {len(country_jobs):,} | Global: {len(global_jobs):,}")

    return country_jobs

# ==========================================
# 6. Processing
# ==========================================
def _work_model(text: str, loc: str) -> str:
    if _REMOTE_RE.search(text) or 'remote' in loc:
        return 'Remote'
    if _HYBRID_RE.search(text):
        return 'Hybrid'
    return 'On-site'

def process_country(country: str, jobs: dict[str, dict]) -> tuple[list[dict], dict[str, int]]:
    counts: dict[str, dict] = {
        level: defaultdict(lambda: {'Total': 0, 'Remote': 0, 'Hybrid': 0, 'On-site': 0})
        for level in ALL_LEVELS
    }
    distribution: dict[str, int] = defaultdict(int)

    for job in jobs.values():
        title = job.get('title', '')
        description = job.get('description', '') or ''
        text = f"{title} {description}"
        level = classify_level(text)
        distribution[level] += 1
        loc = job.get('location', {}).get('display_name', '').lower()
        model = _work_model(text, loc)

        for stack, rx in REGEX_MAP.items():
            if rx.search(text):
                canonical = STACK_ALIASES.get(stack, stack)
                counts[level][canonical]['Total'] += 1
                counts[level][canonical][model] += 1

    data: list[dict] = []
    for level in ALL_LEVELS:
        for stack, info in counts[level].items():
            if info['Total'] > 0:
                data.append({
                    'Country':    COUNTRIES[country]['name'],
                    'Level':      level,
                    'Category':   STACK_TO_CAT[stack],
                    'Technology': stack,
                    'Total':      info['Total'],
                    'Remote':     info['Remote'],
                    'Hybrid':     info['Hybrid'],
                    'On-site':    info['On-site'],
                    '% Remote':   round(info['Remote'] / info['Total'] * 100, 1) if info['Total'] else 0,
                })
    return data, distribution

# ==========================================
# 7. Terminal Report
# ==========================================
def print_report(df: pd.DataFrame, country_name: str) -> None:
    df_country = df[df['Country'] == country_name]
    print(f"\n\n{'#'*60}\n#  MARKET: {country_name.upper()}\n{'#'*60}")
    for level in ALL_LEVELS:
        df_level = df_country[df_country['Level'] == level]
        if df_level.empty: continue
        print(f"\n{'='*60}\n  {level.upper()}\n{'='*60}")
        for cat in CATEGORIES:
            df_cat = df_level[df_level['Category'] == cat].sort_values('Total', ascending=False).head(10)
            if not df_cat.empty:
                print(f"\n--- TOP: {cat.upper()} ---")
                print(df_cat[['Technology', 'Total', 'Remote', 'Hybrid', 'On-site', '% Remote']].to_string(index=False))

# ==========================================
# 8. Excel Exporting
# ==========================================
LEVEL_COLORS = {
    'Senior': 'FF2E75B6', 'Mid-level': 'FF2E86C1', 'Junior': 'FF1E8449',
    'Internship': 'FF8E44AD', 'General': 'FF555555',
}
ALT_ROW_COLOR = 'FFF2F2F2'
THIN_BORDER = Side(style='thin', color='FFCCCCCC')
BORDER_STYLE = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

def _header_style(cell, bg='FF1F3864', fg='FFFFFFFF', bold=True, size=10):
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER_STYLE

def _data_style(cell, alt=False, bold=False, align='center'):
    cell.font = Font(name='Arial', size=9, bold=bold)
    if alt: cell.fill = PatternFill('solid', start_color=ALT_ROW_COLOR)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER_STYLE

def _auto_width(ws, extra=4):
    for col in ws.columns:
        max_len, col_letter = 0, get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            except Exception: pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 50)

def _create_country_sheet(wb: Workbook, df_country: pd.DataFrame, country_name: str):
    ws = wb.create_sheet(country_name)
    ws.freeze_panes = 'A3'
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'Market Analysis — {country_name} (last {DAYS_AGO} days)'
    title_cell.font, title_cell.fill = Font(name='Arial', bold=True, size=13, color='FFFFFFFF'), PatternFill('solid', start_color='FF1F3864')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Level', 'Category', 'Technology', 'Total', 'Remote', 'Hybrid', 'On-site', '% Remote', 'Rank by Category']
    for col, header in enumerate(headers, 1): _header_style(ws.cell(row=2, column=col, value=header))
    ws.row_dimensions[2].height = 22

    df_sorted = df_country.sort_values(['Level', 'Category', 'Total'], ascending=[True, True, False])
    row_idx = 3
    for _, row in df_sorted.iterrows():
        alt = (row_idx % 2 == 0)
        vals = [row['Level'], row['Category'], row['Technology'], row['Total'], row['Remote'], row['Hybrid'], row['On-site'], row['% Remote'], '']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _data_style(cell, alt=alt, align='left' if col in (1, 2, 3) else 'center')
            if col == 1: cell.font = Font(name='Arial', size=9, bold=True, color=LEVEL_COLORS.get(row['Level'], 'FF555555'))
            if col == 8: cell.number_format = '0.0"%"'
        row_idx += 1

    row_idx = 3
    for _, row in df_sorted.iterrows():
        subset = df_sorted[(df_sorted['Level'] == row['Level']) & (df_sorted['Category'] == row['Category'])]
        rank = list(subset['Technology']).index(row['Technology']) + 1
        _data_style(ws.cell(row=row_idx, column=9, value=f"#{rank}"), alt=(row_idx % 2 == 0))
        row_idx += 1

    if row_idx > 3:
        ws.conditional_formatting.add(f'D3:D{row_idx-1}', ColorScaleRule(start_type='min', start_color='FFFFFFFF', end_type='max', end_color='FF1F3864'))
    _auto_width(ws)
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 16, 24, 22

def _create_all_data_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('All Data', 0)
    ws.freeze_panes = 'A3'
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value, title_cell.font, title_cell.fill = 'Complete Analysis — Brazil + USA', Font(name='Arial', bold=True, size=13, color='FFFFFFFF'), PatternFill('solid', start_color='FF1F3864')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Country', 'Level', 'Category', 'Technology', 'Total', 'Remote', 'Hybrid', 'On-site', '% Remote']
    for col, header in enumerate(headers, 1): _header_style(ws.cell(row=2, column=col, value=header))
    ws.row_dimensions[2].height = 22

    df_sorted = df.sort_values(['Country', 'Level', 'Category', 'Total'], ascending=[True, True, True, False])
    row_idx = 3
    for _, row in df_sorted.iterrows():
        alt = (row_idx % 2 == 0)
        for col, val in enumerate([row['Country'], row['Level'], row['Category'], row['Technology'], row['Total'], row['Remote'], row['Hybrid'], row['On-site'], row['% Remote']], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _data_style(cell, alt=alt, align='left' if col <= 4 else 'center')
            if col == 2: cell.font = Font(name='Arial', size=9, bold=True, color=LEVEL_COLORS.get(str(val), 'FF555555'))
            if col == 9: cell.number_format = '0.0"%"'
        row_idx += 1

    if row_idx > 3:
        ws.conditional_formatting.add(f'E3:E{row_idx-1}', ColorScaleRule(start_type='min', start_color='FFFFFFFF', end_type='max', end_color='FF1F3864'))
    _auto_width(ws)
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width, ws.column_dimensions['D'].width = 10, 16, 24, 22

def _create_comparison_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Comparison BR vs USA')
    ws.freeze_panes = 'A3'
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value, title_cell.font, title_cell.fill = 'Comparison BR vs USA', Font(name='Arial', bold=True, size=13, color='FFFFFFFF'), PatternFill('solid', start_color='FF1F3864')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    pivot = df.groupby(['Category', 'Technology', 'Country'])['Total'].sum().unstack('Country').fillna(0).astype(int).reset_index()
    countries = [col for col in pivot.columns if col not in ('Category', 'Technology')]
    pivot['Grand Total'] = pivot[countries].sum(axis=1)
    pivot = pivot.sort_values(['Category', 'Grand Total'], ascending=[True, False])

    headers = ['Category', 'Technology'] + countries + ['Grand Total']
    for col, header in enumerate(headers, 1): _header_style(ws.cell(row=2, column=col, value=header))
    ws.row_dimensions[2].height = 22

    row_idx = 3
    for _, row in pivot.iterrows():
        vals = [row['Category'], row['Technology']] + [row[p] for p in countries] + [row['Grand Total']]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _data_style(cell, alt=(row_idx % 2 == 0), align='left' if col <= 2 else 'center')
            if col == len(vals): cell.font = Font(name='Arial', size=9, bold=True)
        row_idx += 1

    if row_idx > 3:
        ws.conditional_formatting.add(f'C3:C{row_idx-1}', ColorScaleRule(start_type='min', start_color='FFFFFFFF', end_type='max', end_color='FF1A5276'))
        if len(countries) > 1: ws.conditional_formatting.add(f'D3:D{row_idx-1}', ColorScaleRule(start_type='min', start_color='FFFFFFFF', end_type='max', end_color='FF1A5276'))
    _auto_width(ws)
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 26, 22

def _create_executive_summary_sheet(wb: Workbook, df: pd.DataFrame, distribution_by_country: dict):
    ws = wb.create_sheet('Executive Summary', 0)
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value, title_cell.font, title_cell.fill = 'Tech Market Overview', Font(name='Arial', bold=True, size=14, color='FFFFFFFF'), PatternFill('solid', start_color='FF1F3864')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    row_idx = 3
    for country, cfg in COUNTRIES.items():
        name, dist = cfg['name'], distribution_by_country.get(cfg['name'], {})
        total = sum(dist.values())
        ws.merge_cells(f'A{row_idx}:F{row_idx}')
        cell = ws.cell(row=row_idx, column=1, value=f'{name} — {total} collected jobs')
        cell.font, cell.fill = Font(name='Arial', bold=True, size=11, color='FFFFFFFF'), PatternFill('solid', start_color='FF2E75B6')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

        for level in ['Senior', 'Mid-level', 'General', 'Junior', 'Internship']:
            qty = dist.get(level, 0)
            pct = qty / total * 100 if total else 0
            for col, val in enumerate([level, qty, f'{pct:.1f}%'], 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                _data_style(cell, alt=(row_idx % 2 == 0), align='left' if col == 1 else 'center')
                if col == 1: cell.font = Font(name='Arial', size=9, bold=True, color=LEVEL_COLORS.get(level, 'FF333333'))
            row_idx += 1
        row_idx += 1

    row_idx += 1
    ws.merge_cells(f'A{row_idx}:F{row_idx}')
    cell = ws.cell(row=row_idx, column=1, value='Top 10 Technologies per Country')
    cell.font, cell.fill, cell.alignment = Font(name='Arial', bold=True, size=11, color='FFFFFFFF'), PatternFill('solid', start_color='FF1F3864'), Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1

    col_offset = 1
    for cfg in COUNTRIES.values():
        name = cfg['name']
        top10 = df[df['Country'] == name].groupby('Technology')['Total'].sum().nlargest(10).reset_index()
        ws.cell(row=row_idx, column=col_offset, value=name).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=row_idx, column=col_offset+1, value='Jobs').font = Font(name='Arial', bold=True, size=10)
        for i, (_, row) in enumerate(top10.iterrows(), start=1):
            ws.cell(row=row_idx+i, column=col_offset, value=row['Technology'])
            ws.cell(row=row_idx+i, column=col_offset+1, value=row['Total'])
        col_offset += 3
    _auto_width(ws)

def export_to_excel(df: pd.DataFrame, filename: str, distribution_by_country: dict) -> None:
    wb = Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    _create_executive_summary_sheet(wb, df, distribution_by_country)
    _create_all_data_sheet(wb, df)
    for country, cfg in COUNTRIES.items():
        df_country = df[df['Country'] == cfg['name']]
        if not df_country.empty: _create_country_sheet(wb, df_country, cfg['name'])
    _create_comparison_sheet(wb, df)
    wb.save(filename)
    print(f"\n[+] File exported: {filename}\n    Sheets: {' | '.join(wb.sheetnames)}")

# ==========================================
# 9. Main
# ==========================================
def analyze_job_market() -> None:
    print("=" * 60)
    print("ADVANCED JOB EXTRACTION — BR + USA")
    print(f"Global limit: {JOB_LIMIT:,} unique jobs")
    print("=" * 60)

    all_data: list[dict] = []
    distribution_by_country: dict[str, dict] = {}
    global_jobs: dict[str, dict] = {}

    for country, cfg in COUNTRIES.items():
        if len(global_jobs) >= JOB_LIMIT:
            print(f"\n[!] Limit reached. Skipping {cfg['name']}.")
            break

        print(f"\n\n{'─'*60}\nCollecting jobs: {cfg['name']} ({country.upper()})\nGlobal jobs: {len(global_jobs):,} / {JOB_LIMIT:,}\n{'─'*60}")
        country_jobs = get_country_jobs(country, cfg['terms'], global_jobs)
        data, distribution = process_country(country, country_jobs)
        all_data.extend(data)
        distribution_by_country[cfg['name']] = distribution

        total = sum(distribution.values())
        print(f"\nDistribution by level — {cfg['name']}:")
        for level, qty in sorted(distribution.items(), key=lambda x: -x[1]):
            print(f"   {level:<20} {qty:>6,} jobs ({(qty/total*100 if total else 0):.1f}%)")

    print(f"\n{'='*60}\nTotal globally collected: {len(global_jobs):,} unique jobs\n{'='*60}")

    df = pd.DataFrame(all_data)
    if df.empty:
        print("\nNo data found.")
        return

    for cfg in COUNTRIES.values():
        print_report(df, cfg['name'])
    export_to_excel(df, 'market_analysis_br_usa.xlsx', distribution_by_country)

if __name__ == "__main__":
    analyze_job_market()
